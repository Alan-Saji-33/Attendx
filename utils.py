

import os
from datetime import datetime
from PIL import Image
import customtkinter as ctk
from config import IMAGES_PATH, ENCODINGS_PATH, REPORTS_PATH


def ensure_directories():
    """Create necessary directories if they don't exist"""
    directories = [IMAGES_PATH, ENCODINGS_PATH, REPORTS_PATH]
    for directory in directories:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"INFO: Created directory: {directory}")


def get_current_date():
    """Get current date in YYYY-MM-DD format"""
    return datetime.now().strftime("%Y-%m-%d")


def get_current_time():
    """Get current time in HH:MM:SS format"""
    return datetime.now().strftime("%H:%M:%S")


def get_current_datetime():
    """Get current datetime formatted"""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_display_date():
    """Get display-friendly date format"""
    return datetime.now().strftime("%A, %B %d, %Y")


def get_display_time():
    """Get display-friendly time format"""
    return datetime.now().strftime("%I:%M:%S %p")


def format_date(date_str):
    """Format date string to display format"""
    try:
        date_obj = datetime.strptime(str(date_str), "%Y-%m-%d")
        return date_obj.strftime("%b %d, %Y")
    except:
        return date_str


def format_time(time_str):
    """Format time string to display format"""
    try:
        time_obj = datetime.strptime(str(time_str), "%H:%M:%S")
        return time_obj.strftime("%I:%M %p")
    except:
        return time_str


def generate_student_id():
    """Generate unique student ID"""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"STU{timestamp}"


def save_image(image_array, filename):
    """Save image array to file"""
    try:
        ensure_directories()
        filepath = os.path.join(IMAGES_PATH, filename)
        Image.fromarray(image_array).save(filepath)
        return filepath
    except Exception as e:
        print(f"ERROR: Error saving image: {e}")
        return None


def resize_image(image_path, size=(200, 200)):
    """Resize image to specified size"""
    try:
        img = Image.open(image_path)
        img = img.resize(size, Image.Resampling.LANCZOS)
        return img
    except Exception as e:
        print(f"ERROR: Error resizing image: {e}")
        return None


def show_toast(parent, message, message_type="success"):
    """Show toast notification"""
    colors = {
        "success": "#10b981",
        "error": "#ef4444",
        "info": "#06b6d4",
        "warning": "#f59e0b"
    }

    icons = {
        "success": "✅",
        "error": "❌",
        "info": "ℹ️",
        "warning": "⚠️"
    }

    color = colors.get(message_type, colors["info"])
    icon = icons.get(message_type, icons["info"])

    # Create toast frame
    toast = ctk.CTkFrame(
        parent,
        fg_color=color,
        corner_radius=10
    )
    toast.place(relx=0.5, rely=0.1, anchor="center")

    # Toast label
    label = ctk.CTkLabel(
        toast,
        text=f"{icon}  {message}",
        font=("Roboto", 14, "bold"),
        text_color="white",
        padx=20,
        pady=10
    )
    label.pack()

    # Auto-hide after 3 seconds
    parent.after(3000, lambda: toast.destroy())


def validate_email(email):
    """Validate email format"""
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_phone(phone):
    """Validate phone number format"""
    import re
    pattern = r'^[0-9]{10}$'
    return re.match(pattern, phone) is not None


def center_window(window, width, height):
    """Center window on screen"""
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")


def clear_frame(frame):
    """Clear all widgets from frame"""
    for widget in frame.winfo_children():
        widget.destroy()


def export_to_excel(data, filename, headers):
    """Export data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        ensure_directories()
        filepath = os.path.join(REPORTS_PATH, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header style
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        return filepath
    except Exception as e:
        print(f"ERROR: Error exporting to Excel: {e}")
        return None


def create_hover_effect(widget, hover_color, normal_color):
    """Add hover effect to widget"""
    def on_enter(e):
        widget.configure(fg_color=hover_color)

    def on_leave(e):
        widget.configure(fg_color=normal_color)

    widget.bind("<Enter>", on_enter)
    widget.bind("<Leave>", on_leave)


class LoadingDialog(ctk.CTkToplevel):
    """Loading dialog with progress indication"""

    def __init__(self, parent, title="Processing", message="Please wait..."):
        super().__init__(parent)

        self.title(title)
        self.geometry("400x150")
        self.resizable(False, False)

        # Center on parent
        self.transient(parent)
        self.grab_set()

        # Main frame
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Message label
        self.message_label = ctk.CTkLabel(
            main_frame,
            text=message,
            font=("Roboto", 16),
            wraplength=350
        )
        self.message_label.pack(pady=(0, 15))

        # Progress bar
        self.progress = ctk.CTkProgressBar(
            main_frame,
            width=360,
            height=10,
            mode="indeterminate"
        )
        self.progress.pack()
        self.progress.start()

        # Center on screen
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    def update_message(self, message):
        """Update loading message"""
        self.message_label.configure(text=message)
        self.update()

    def close(self):
        """Close loading dialog"""
        self.progress.stop()
        self.destroy()

def send_attendance_email(student_name, student_email, date, period, status, marked_by):
    """Send email notification to student for absence"""
    from config import SMTP_CONFIG
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    if not student_email:
        print(f"⚠️ No email for {student_name}, skipping notification.")
        return False

    try:
        msg = MIMEMultipart()
        msg['From'] = f"{SMTP_CONFIG['from_name']} <{SMTP_CONFIG['username']}>"
        msg['To'] = student_email
        msg['Subject'] = f"Absence Notification - Period {period} ({date})"

        html_body = f"""
        <html>
        <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
            <div style="max-width: 600px; margin: 20px auto; border: 1px solid #e1e4e8; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.05);">
                <div style="background: linear-gradient(135deg, #2563eb 0%, #7c3aed 100%); padding: 30px; text-align: center; color: white;">
                    <h1 style="margin: 0; font-size: 24px; letter-spacing: 1px;">Attendance Alert</h1>
                </div>
                <div style="padding: 30px; background-color: white;">
                    <p style="font-size: 18px; margin-bottom: 20px;">Hello <strong>{student_name}</strong>,</p>
                    <p style="margin-bottom: 20px;">This is an automated notification from the <strong>{SMTP_CONFIG['from_name']}</strong> regarding your attendance today.</p>
                    
                    <div style="background-color: #f8fafc; border-left: 4px solid #ef4444; padding: 20px; border-radius: 6px; margin-bottom: 25px;">
                        <h3 style="margin-top: 0; color: #ef4444; font-size: 16px; text-transform: uppercase;">Absence Recorded</h3>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 8px 0; color: #64748b; width: 120px;">Date:</td>
                                <td style="padding: 8px 0; font-weight: 600;">{date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b;">Period:</td>
                                <td style="padding: 8px 0; font-weight: 600;">{period}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b;">Marked By:</td>
                                <td style="padding: 8px 0; font-weight: 600;">{marked_by}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b;">Status:</td>
                                <td style="padding: 8px 0;"><span style="background-color: #fee2e2; color: #ef4444; padding: 4px 10px; border-radius: 20px; font-size: 12px; font-weight: 700;">{status.upper()}</span></td>
                            </tr>
                        </table>
                    </div>

                    <p style="margin-bottom: 25px;">If you were present in the class or have a valid reason for your absence, please log in to your dashboard to submit a <strong>Claim</strong> for approval.</p>
                    
                    <div style="text-align: center; margin-bottom: 20px;">
                        <a href="#" style="background: linear-gradient(135deg, #2563eb 0%, #7c3aed 100%); color: white; padding: 14px 28px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block; box-shadow: 0 4px 10px rgba(37, 99, 235, 0.3);">View Student Dashboard</a>
                    </div>
                </div>
                <div style="padding: 20px; background-color: #f1f5f9; text-align: center; font-size: 12px; color: #94a3b8;">
                    <p style="margin: 0;">This is an automated system email. Please do not reply directly to this message.</p>
                    <p style="margin: 5px 0 0 0;">&copy; 2026 {SMTP_CONFIG['from_name']}</p>
                </div>
            </div>
        </body>
        </html>
        """
        msg.attach(MIMEText(html_body, 'html'))

        # Connect and send
        server = smtplib.SMTP(SMTP_CONFIG['server'], SMTP_CONFIG['port'])
        server.starttls()
        server.login(SMTP_CONFIG['username'], SMTP_CONFIG['password'])
        server.send_message(msg)
        server.quit()
        
        print(f"SUCCESS: Notification sent to {student_name} ({student_email})")
        return True
    except Exception as e:
        print(f"ERROR: Failed to send email to {student_email}: {e}")
        return False
